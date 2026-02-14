
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define the data structure
features = [
    {"Feature": "Feature 0: UI & Hardware", "Type": "Microphone Input", 
     "English": "Hello, how are you?", 
     "Hindi": "Namaste, kaise hain aap? (नमस्ते, कैसे हैं आप?)", 
     "Bengali": "Nomoshkar, kemon achen? (নমস্কার, কেমন আছেন?)",
     "Kannada": "Namaskara, hegiddira? (ನಮಸ್ಕಾರ, ಹೇಗಿದ್ದೀರಾ?)"},
    
    {"Feature": "Feature 0: UI & Hardware", "Type": "Action Bar: Copy", 
     "English": "Test Copy Button", 
     "Hindi": "Test Copy Button", 
     "Bengali": "Test Copy Button",
     "Kannada": "Test Copy Button"},

    {"Feature": "Feature 0: UI & Hardware", "Type": "Action Bar: Save", 
     "English": "Test Save Button", 
     "Hindi": "Test Save Button", 
     "Bengali": "Test Save Button",
     "Kannada": "Test Save Button"},
     
    {"Feature": "Feature 0: UI & Hardware", "Type": "Action Bar: PDF Download", 
     "English": "Test PDF Download", 
     "Hindi": "Test PDF Download", 
     "Bengali": "Test PDF Download",
     "Kannada": "Test PDF Download"},

    {"Feature": "Feature 1: Lesson Plan", "Type": "Science (Class 5)", 
     "English": "Explain the process of Photosynthesis for Class 5 students in a rural primary school.", 
     "Hindi": "ग्रामीण प्राथमिक विद्यालय के कक्षा 5 के छात्रों के लिए प्रकाश संश्लेषण (Photosynthesis) की प्रक्रिया को समझाएं।", 
     "Bengali": "গ্রামের প্রাথমিক স্কুলের পঞ্চম শ্রেণীর পড়ুয়াদের জন্য সালোকসংশ্লেষণ (Photosynthesis) প্রক্রিয়াটি বুঝিয়ে বলুন।",
     "Kannada": "ಗ್ರಾಮೀಣ ಪ್ರಾಥಮಿಕ ಶಾಲೆಯ 5 ನೇ ತರಗತಿ ವಿದ್ಯಾರ್ಥಿಗಳಿಗೆ ದ್ಯುತಿಸಂಶ್ಲೇಷಣೆ ಪ್ರಕ್ರಿಯೆಯನ್ನು ವಿವರಿಸಿ."},

    {"Feature": "Feature 1: Lesson Plan", "Type": "Math (Class 3)", 
     "English": "Teach the concept of 'Money and Market' to Class 3 using only chalk and common household items.", 
     "Hindi": "केवल चॉक और सामान्य घरेलू वस्तुओं का उपयोग करके कक्षा 3 को 'पैसा और बाजार' (Money and Market) की अवधारणा सिखाएं।", 
     "Bengali": "শুধুমাত্র চক আর সাধারণ গৃহস্থালি জিনিস ব্যবহার করে তৃতীয় শ্রেণীর ছাত্রছাত্রীদের 'টাকা আর বাজার'-এর ধারণাটি শেখান।",
     "Kannada": "ಕೇವಲ ಚಾಕ್ ಮತ್ತು ಸಾಮಾನ್ಯ ಗೃಹೋಪಯೋಗಿ ವಸ್ತುಗಳನ್ನು ಬಳಸಿ 3 ನೇ ತರಗತಿಗೆ 'ಹಣ ಮತ್ತು ಮಾರುಕಟ್ಟೆ' ಪರಿಕಲ್ಪನೆಯನ್ನು ಬೋಧಿಸಿ."},

    {"Feature": "Feature 1: Lesson Plan", "Type": "Geography (Class 9)", 
     "English": "A 45-minute lesson on 'The Indian Monsoon' focusing on why it matters to farmers in Maharashtra.", 
     "Hindi": "महर्षि के किसानों के लिए 'भारतीय मानसून' क्यों महत्वपूर्ण है, इस पर केंद्रित 45 मिनट का पाठ तैयार करें।", 
     "Bengali": "মহারাষ্ট্রের কৃষকদের কাছে 'ভারতীয় মৌসুমি বায়ু' বা মনসুন কেন এত জরুরি, তার ওপর জোর দিয়ে ৪৫ মিনিটের একটি পাঠ তৈরি করুন।",
     "Kannada": "ಭಾರತೀಯ ಮಾನ್ಸೂನ್ ಕುರಿತು 45 ನಿಮಿಷಗಳ ಪಾಠ, ಇದು ಮಹಾರಾಷ್ಟ್ರದ ರೈತರಿಗೆ ಹೇಗೆ ಮುಖ್ಯವಾಗಿದೆ ಎಂಬುವುದರ ಮೇಲೆ ಕೇಂದ್ರೀಕರಿಸಿ."},

    {"Feature": "Feature 1: Lesson Plan", "Type": "Social Studies (Class 8)", 
     "English": "The story of India's Independence highlighting localized heroes from different regions.", 
     "Hindi": "विभिन्न क्षेत्रों के स्थानीय नायकों को उजागर करते हुए भारत की स्वतंत्रता की कहानी।", 
     "Bengali": "বিভিন্ন অঞ্চলের স্থানীয় বিপ্লবীদের কথা তুলে ধরে ভারতের স্বাধীনতা সংগ্রামের কাহিনী।",
     "Kannada": "ವಿವಿಧ ಪ್ರದೇಶಗಳ ಸ್ಥಳೀಯ ವೀರರನ್ನು ಎತ್ತಿ ತೋರಿಸುವ ಭಾರತದ ಸ್ವಾತಂತ್ರ್ಯದ ಕಥೆ."},

    {"Feature": "Feature 1: Lesson Plan", "Type": "Multi-Grade (Class 1-2)", 
     "English": "A combined lesson on 'Cleanliness and Hygiene' for a single classroom.", 
     "Hindi": "एक ही कक्षा के लिए 'सफाई और स्वच्छता' पर एक संयुक्त पाठ।", 
     "Bengali": "একই শ্রেণীকক্ষের জন্য 'পরিচ্ছন্নতা ও স্বাস্থ্যবিধি' নিয়ে একটি যৌথ বা কম্বাইন্ড পাঠ (Combined Lesson)।",
     "Kannada": "ಒಂದೇ ತರಗತಿಯಲ್ಲಿ 'ಶುಚಿತ್ವ ಮತ್ತು ನೈರ್ಮಲ್ಯ' ಕುರಿತು ಸಂಯೋಜಿತ ಪಾಠ."},

    {"Feature": "Feature 2: Visual Aid", "Type": "Biology (Class 10)", 
     "English": "Draw a detailed, labeled diagram of a Hibiscus flower (Gudhal) showing the stigma, petals, and ovary.", 
     "Hindi": "गुड़हल (Hibiscus) के फूल का एक विस्तृत, नामांकित चित्र बनाएं जिसमें वर्तिकाग्र (stigma), पंखुड़ियाँ (petals) और अंडाशय (ovary) दिखाई दें।", 
     "Bengali": "গর্ভমুণ্ড (Stigma), পাপড়ি আর ডিম্বাশয় (Ovary) দেখিয়ে একটি জবা ফুলের বিস্তারিত চিহ্নিত চিত্র আঁকুন।",
     "Kannada": "ಶಲಾಕಾಗ್ರ, ದಳಗಳು ಮತ್ತು ಅಂಡಾಶಯವನ್ನು ತೋರಿಸುವ ದಾಸವಾಳ ಹೂವಿನ ವಿವರವಾದ, ಹೆಸರಿಸಲಾದ ರೇಖಾಚಿತ್ರವನ್ನು ಬಿಡಿಸಿ."},

    {"Feature": "Feature 2: Visual Aid", "Type": "Physics (Class 8)", 
     "English": "Create a simple blackboard diagram showing how a hand-pump (Boring) works to pull water from the ground.", 
     "Hindi": "जमीन से पानी खींचने के लिए हैंड-पंप (बोरिंग) कैसे काम करता है, यह दिखाने वाला एक सरल ब्लैकबोर्ड चित्र बनाएं।", 
     "Bengali": "মাটির নিচ থেকে জল তুলতে হ্যান্ডপাম্প বা টিউবওয়েল কীভাবে কাজ করে, তা বোঝানোর জন্য একটি সরল ব্ল্যাকবোর্ড ডায়াগ্রাম তৈরি করুন।",
     "Kannada": "ಭೂಮಿಯಿಂದ ನೀರನ್ನು ತೆಗೆಯಲು ಕೈ ಪಂಪ್ (ಬೋರಿಂಗ್) ಹೇಗೆ ಕೆಲಸ ಮಾಡುತ್ತದೆ ಎಂಬುದನ್ನು ತೋರಿಸುವ ಸರಳ ಕಪ್ಪುಹಲಗೆಯ ಚಿತ್ರವನ್ನು ರಚಿಸಿ."},

    {"Feature": "Feature 2: Visual Aid", "Type": "Astronomy (Class 6)", 
     "English": "A sketch of the Solar System showing the relative positions of the Earth, Moon, and Sun during a Lunar Eclipse.", 
     "Hindi": "चंद्र ग्रहण के दौरान पृथ्वी, चंद्रमा और सूर्य की सापेक्ष स्थिति को दिखाते हुए सौर मंडल का एक स्केच की रूपरेखा तैयार करें।", 
     "Bengali": "চন্দ্রগ্রহণের সময় পৃথিবী, চাঁদ আর সূর্যের অবস্থান দেখিয়ে সৌরজগতের একটি স্কেচ বা রেখাচিত্র।",
     "Kannada": "ಚಂದ್ರಗ್ರಹಣದ ಸಮಯದಲ್ಲಿ ಭೂಮಿ, ಚಂದ್ರ ಮತ್ತು ಸೂರ್ಯನ ಸಾಪೇಕ್ಷ ಸ್ಥಾನಗಳನ್ನು ತೋರಿಸುವ ಸೌರವ್ಯೂಹದ ರೇಖಾಚಿತ್ರ."},

    {"Feature": "Feature 3: Quiz Generator", "Type": "History (Class 7)", 
     "English": "5 Multiple Choice Questions on the 'Life of Mahatma Gandhi'.", 
     "Hindi": "'महात्मा गांधी के जीवन' पर 5 बहुविकल्पीय प्रश्न।", 
     "Bengali": "'মহাত্মা গান্ধীর জীবন' নিয়ে ৫টি মাল্টিপল চয়েস প্রশ্ন (MCQ)।",
     "Kannada": "'ಮಹಾತ್ಮ ಗಾಂಧಿಯವರ ಜೀವನ' ಬಗ್ಗೆ ಕನ್ನಡದಲ್ಲಿ 5 ಬಹು ಆಯ್ಕೆಯ ಪ್ರಶ್ನೆಗಳು."},

    {"Feature": "Feature 3: Quiz Generator", "Type": "EVS (Class 4)", 
     "English": "A 10-question True/False quiz about 'Saving Water' with detailed explanations.", 
     "Hindi": "विस्तृत स्पष्टीकरण के साथ 'जल संरक्षण' (Saving Water) के बारे में 10-प्रश्न वाला सही/गलत प्रश्नोत्तरी।", 
     "Bengali": "'জল সংরক্ষণ'-এর ওপর বিস্তারিত ব্যাখ্যাসহ ১০টি ঠিক/ভুল প্রশ্ন।",
     "Kannada": "ನೀರು ಉಳಿಸುವ ಬಗ್ಗೆ ವಿವರವಾದ ವಿವರಣೆಗಳೊಂದಿಗೆ 10 ಪ್ರಶ್ನೆಗಳ ಹೌದು/ಅಲ್ಲ ರಸಪ್ರಶ್ನೆ."},

    {"Feature": "Feature 4: Teacher Training", "Type": "Classroom Management", 
     "English": "How can I handle a student who is constantly talking while I am writing on the blackboard?", 
     "Hindi": "मैं उस छात्र को कैसे संभालूँ जो मेरे ब्लैकबोर्ड पर लिखते समय लगातार बात कर रहा है?", 
     "Bengali": "আমি যখন ব্ল্যাকবোর্ডে লিখছি তখন যে ছাত্রটি অনবরত কথা বলছে, তাকে আমি কীভাবে সামলাব?",
     "Kannada": "ನಾನು ಕಪ್ಪುಹಲಗೆಯ ಮೇಲೆ ಬರೆಯುವಾಗ ನಿರಂತರವಾಗಿ ಮಾತನಾಡುತ್ತಿರುವ ವಿದ್ಯಾರ್ಥಿಯನ್ನು ನಾನು ಹೇಗೆ ನಿಭಾಯಿಸಬಹುದು?"},

    {"Feature": "Feature 4: Teacher Training", "Type": "Inclusion", 
     "English": "How can I make my Science lesson more inclusive for a student with a mild hearing impairment?", 
     "Hindi": "हल्की श्रवण बाधित छात्र के लिए मैं अपने विज्ञान के पाठ को और अधिक समावेशी (inclusive) कैसे बना सकता हूँ?", 
     "Bengali": "শ্রবণশক্তি কিছুটা কম এমন একজন ছাত্রের জন্য আমি আমার বিজ্ঞানের ক্লাসকে আরও অন্তর্ভুক্তিমূলক (Inclusive) করব কীভাবে?",
     "Kannada": "ಕಡಿಮೆ ಕೇಳುವ ದೋಷವಿರುವ ವಿದ್ಯಾರ್ಥಿಗಾಗಿ ನನ್ನ ವಿಜ್ಞಾನ ಪಾಠವನ್ನು ಹೆಚ್ಚು ಒಳಗೊಳ್ಳುವಂತೆ (inclusive) ಮಾಡುವುದು ಹೇಗೆ?"},

    {"Feature": "Feature 5: Worksheet Wizard", "Type": "Comprehension", 
     "English": "Create 5 fill-in-the-blank questions based on the uploaded textbook page about 'The Sun and Planets'.", 
     "Hindi": "'सूर्य और ग्रह' के बारे में पाठ्यपुस्तक के पृष्ठ के आधार पर 5 रिक्त स्थान भरने वाले प्रश्न बनाएं।", 
     "Bengali": "'সূর্য ও তার গ্রহ' বিষয়ক টেক্সটবুক পৃষ্ঠা থেকে ৫টি শূন্যস্থান পূরণ (Fill in the blanks) প্রশ্ন তৈরি করুন।",
     "Kannada": "ಅಪ್-ಲೋಡ್ ಮಾಡಿದ 'ಸೂರ್ಯ ಮತ್ತು ಗ್ರಹಗಳು' ಎಂಬ ಪಠ್ಯಪುಸ್ತಕದ ಪುಟದ ಆಧಾರದ ಮೇಲೆ 5 ಖಾಲಿ ಜಾಗಗಳನ್ನು ತುಂಬುವ ಪ್ರಶ್ನೆಗಳನ್ನು ರಚಿಸಿ."},

    {"Feature": "Feature 5: Worksheet Wizard", "Type": "Math Word Problems", 
     "English": "Extract the 3 word problems from this page and create 2 more similar ones for extra practice.", 
     "Hindi": "इस पृष्ठ से 3 शब्द समस्याओं (word problems) को निकालें और अतिरिक्त अभ्यास के लिए 2 और समान समस्याएं बनाएं।", 
     "Bengali": "এই পৃষ্ঠা থেকে ৩টি গাণিতিক সমস্যা (Word Problems) বের করুন এবং অনুশীলনের জন্য আরও ২টি একই ধরণের সমস্যা বানিয়ে দিন।",
     "Kannada": "ಈ ಪುಟದಿಂದ 3 ಪದ ಸಮಸ್ಯೆಗಳನ್ನು (word problems) ಹೊರತೆಗೆಯಿರಿ ಮತ್ತು ಹೆಚ್ಚಿನ ಅಭ್ಯಾಸಕ್ಕಾಗಿ ಅಂತಹದ್ದೇ 2 ಸಮಸ್ಯೆಗಳನ್ನು ರಚಿಸಿ."},

    {"Feature": "Feature 6: Rubric Generator", "Type": "Art Project", 
     "English": "Create a rubric for a Class 5 project where they have to make a model of a 'Village Market' using recycled materials.", 
     "Hindi": "कक्षा 5 की परियोजना के लिए एक रूब्रिक बनाएं जहाँ उन्हें पुनर्नवीनीकरण सामग्री (recycled materials) का उपयोग करके 'गाँव के बाज़ार' का मॉडल बनाना है।", 
     "Bengali": "ফেলে দেওয়া জিনিস বা রিসাইকেলড মেটিরিয়াল দিয়ে 'গ্রামের হাট'-এর মডেল বানানোর ক্লাস ফাইভের প্রজেক্টের জন্য একটি রুব্রিক (Rubric) তৈরি করুন।",
     "Kannada": "ಮರುಬಳಕೆ ಮಾಡಲಾದ ವಸ್ತುಗಳನ್ನು ಬಳಸಿ 'ಗ್ರಾಮದ ಮಾರುಕಟ್ಟೆ'ಯ ಮಾದರಿಯನ್ನು ತಯಾರಿಸಬೇಕಾದ 5 ನೇ ತರಗತಿಯ ಪ್ರಾಜೆಕ್ಟ್-ಗಾಗಿ ರಬ್ರಿಕ್ ಅನ್ನು ರಚಿಸಿ."},

    {"Feature": "Feature 6: Rubric Generator", "Type": "Essay Writing", 
     "English": "A simple rubric for a 100-word essay on 'Importance of Forests' for Class 6.", 
     "Hindi": "कक्षा 6 के लिए 'वनों का महत्व' पर 100-शब्दों के निबंध के लिए एक सरल रूब्रिक।", 
     "Bengali": "ষষ্ঠ শ্রেণীর জন্য 'বনজঙ্গলের গুরুত্ব' নিয়ে ১০০ শব্দের প্রবন্ধের একটি সহজ রুব্রিক।",
     "Kannada": "6 ನೇ ತರಗತಿಗೆ 'ಅರಣ್ಯಗಳ ಮಹತ್ವ' ಕುರಿತ 100 ಪದಗಳ ಪ್ರಬಂಧಕ್ಕಾಗಿ ಸರಳ ರಬ್ರಿಕ್."},

    {"Feature": "Feature 7: Virtual Field Trip", "Type": "Heritage", 
     "English": "Take my students on a tour of the 'Ajanta and Ellora Caves' to see the ancient rock-cut architecture.", 
     "Hindi": "प्राचीन रॉक-कट वास्तुकला को देखने के लिए मेरे छात्रों को 'अजंता और एलोरा की गुफाओं' की सैर पर ले जाएं।", 
     "Bengali": "প্রাচীন প্রস্তরশিল্প বা রক-কাট আর্কিটেকচার দেখাতে আমার ছাত্রছাত্রীদের 'অজন্তা ও ইলোরা গুহা' সফরে নিয়ে চলুন।",
     "Kannada": "ಪ್ರಾಚೀನ ಕಲ್ಲಿನಲ್ಲಿ ಕೆತ್ತಲಾದ ವಾಸ್ತುಶಿಲ್ಪವನ್ನು ನೋಡಲು ನನ್ನ ವಿದ್ಯಾರ್ಥಿಗಳನ್ನು 'ಅಜಂತಾ ಮತ್ತು ಎಲ್ಲೋರಾ ಗುಹೆಗಳ' ಪ್ರವಾಸಕ್ಕೆ ಕರೆದೊಯ್ಯಿರಿ."},

    {"Feature": "Feature 7: Virtual Field Trip", "Type": "Nature", 
     "English": "A trip to the 'Sundarbans' to learn about the mangrove ecosystem and the Bengal Tiger.", 
     "Hindi": "मैंग्रोव पारिस्थितिकी तंत्र (ecosystem) और बंगाल टाइगर के बारे में जानने के लिए 'सुंदरवन' की यात्रा।", 
     "Bengali": "ম্যানগ্রোভ ইকোসিস্টেম আর রয়্যাল বেঙ্গল টাইগার সম্পর্কে জানতে 'সুন্দরবন'-এ একটি ভার্চুয়াল ট্যুর।",
     "Kannada": "ಮ್ಯಾಂಗ್ರೋವ್ ಪರಿಸರ ವ್ಯವಸ್ಥೆ ಮತ್ತು ಬಂಗಾಳದ ಹುಲಿ ಬಗ್ಗೆ ತಿಳಿಯಲು 'ಸುಂದರಬನ' ಪ್ರವಾಸ."},

    {"Feature": "Feature 8: Instant Answer", "Type": "Current Affairs", 
     "English": "What is the significance of the Republic Day parade in India for a 10-year-old?", 
     "Hindi": "एक 10 साल के बच्चे के लिए भारत में गणतंत्र दिवस परेड का क्या महत्व है?", 
     "Bengali": "১০ বছরের বাচ্চার কাছে ভারতের প্রজাতন্ত্র দিবসের কুচকাওয়াজ বা প্যারেড-এর গুরুত্ব কী?",
     "Kannada": "10 ವರ್ಷದ ಮಗುವಿಗೆ ಭಾರತದಲ್ಲಿ ಗಣರಾಜ್ಯೋತ್ಸವ ಪರೇಡ್-ನ ಮಹತ್ವವೇನು?"},

    {"Feature": "Feature 8: Instant Answer", "Type": "Science", 
     "English": "Why do we see rainbows after rain? Explain with a simple analogy.", 
     "Hindi": "बारिश के बाद हमें इंद्रधनुष क्यों दिखाई देता है? एक सरल सादृश्य (analogy) के साथ समझाएं।", 
     "Bengali": "বৃষ্টির পরে আমরা রামধনু (Rainbow) দেখি কেন? সহজ উদাহরণ দিয়ে বুঝিয়ে বলুন।",
     "Kannada": "ಮಳೆಯ ನಂತರ ನಾವು ಕಾಮನಬಿಲ್ಲನ್ನು ಏಕೆ ನೋಡುತ್ತೇವೆ? ಸರಳ ಉದಾಹರಣೆಯೊಂದಿಗೆ ವಿವರಿಸಿ."}
]

# Create Workbook
wb = Workbook()
# Remove default sheet
wb.remove(wb.active)

# Helper function to create sheets
def create_tracker_sheet(sheet_name, lang_key):
    ws = wb.create_sheet(sheet_name)
    
    # Headers
    headers = ["Feature", "Test Scenario", "Input Prompt (" + sheet_name + ")", "Status (Pass/Fail)", "Observations / Bugs", "Output Quality (1-5)"]
    ws.append(headers)

    # Style Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A") # Dark Blue
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Add Data
    for item in features:
        row = [
            item["Feature"],
            item["Type"],
            item[lang_key],  # Language specific prompt
            "", # Status
            "", # Observations
            ""  # Quality
        ]
        ws.append(row)
    
    # Formatting Columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 60 # Prompt is wide
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 15

    # Align Cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))


# Create Sheets for each language
create_tracker_sheet("English Test", "English")
create_tracker_sheet("Hindi Test", "Hindi")
create_tracker_sheet("Bengali Test", "Bengali")
create_tracker_sheet("Kannada Test", "Kannada")

# Metadata Sheet
ws_meta = wb.create_sheet("Testing Metadata")
ws_meta['A1'] = "Project"
ws_meta['B1'] = "SahayakAI"
ws_meta['A2'] = "Date"
ws_meta['B2'] = "February 6, 2026"
ws_meta['A3'] = "Objective"
ws_meta['B3'] = "Multilingual QA Tracking for 8 Core Pedagogical Features + UI Tests"

# Save
file_path = "/Users/sargupta/SahayakAIV2/sahayakai/outputs/investment_and_proposals/SahayakAI_QA_Tracker.xlsx"
wb.save(file_path)
print(f"Created QA Tracker at: {file_path}")
